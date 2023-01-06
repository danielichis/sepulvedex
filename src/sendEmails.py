import smtplib
from email.message import EmailMessage
import ssl
from pathlib import Path
import os
import sys
from utilities import pathsManager as pm
from openpyxl import load_workbook

class sendEmails:
    def __init__(self):
        self.msg = EmailMessage()
        self.msg['Subject'] = 'Kardexs confrontados'        
        self.msg['From'] = 'Bot Sepúlveda'
        self.configPath = os.path.join(pm().currentFolderPath, 'config.xlsx')
        self.wb = load_workbook(self.configPath)
        self.ws = self.wb['conf']
        self.correo = self.ws['H2'].value

        self.msg['To'] = self.correo


    def send(self):
        print('ENVIANDO CORREO, SEA PACIENTE.')
        directory = os.path.join(pm().currentFolderPath, 'KardexsOut')
        files = os.listdir(directory)
        files = [f for f in files if f.endswith('.docx')]
        
        for file in files:
            docxPath = os.path.join(directory, file)
            with open(docxPath, 'rb') as f:
                file_data = f.read()
                file_name = f.name
                self.msg.add_attachment(file_data, maintype = 'application', subtype = 'docx', filename = file_name)

        context = ssl.create_default_context()

        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context= context) as smtp:
            smtp.login("sepulvedabot0@gmail.com", "fwmjcpowfrjiautv")
            smtp.send_message(self.msg)
            smtp.quit()
        print(f'LISTO! CORREO ENVIADO A {self.correo}, REVISE POR FAVOR.')
       


if __name__=='__main__':
    x = sendEmails()
    x.send()