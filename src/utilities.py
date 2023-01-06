import openpyxl
import pandas as pd
import os
import sys

#import PATH from os
from pathlib import Path
class pathsManager:
    def __init__(self):
        self.application_path2=None
        self.currentFolderPath=Path(os.getcwd())
        self.get_application_path()
    def get_application_path(self):
    # determine if application is a script file or frozen exe
        if getattr(sys, 'frozen', False):
            self.application_path2 = os.path.dirname(sys.executable)
        elif __file__:
            self.application_path2 = os.path.dirname(__file__)
        self.currentFolderPath = Path(self.application_path2).parent
        
class configData:
    def __init__(self,configPath) -> None:
        self.wordPath=configPath
        self.data=None
        self.df=self.get_data_config()
        self.pm=None
    def get_data_config(self):
        self.pm=pathsManager()
        #metodo para sacar el rango de un excel a partir de su ruta
        configfile=openpyxl.load_workbook(os.path.join(self.pm.currentFolderPath,"config.xlsx"))
        sheet=configfile.worksheets[0]
        maxr=sheet.max_row
        maxc=sheet.max_column
        maxcLetter=self.get_column_letter(maxc)
        rangeData=f"A1:{maxcLetter}{maxr}"
        self.data=[]
        headersConfig=[]
        for i,row in enumerate(sheet[rangeData]):
            rowdata=[]
            if i==0:    
                for cell in row:
                    headersConfig.append(cell.value)
            else:
                for cell in row:
                    if cell.value is None:
                        rowdata.append("")
                    else:
                        rowdata.append(cell.value)
                self.data.append(rowdata)
        df=pd.DataFrame(self.data,columns=headersConfig)
        return df
    def get_column_letter(self,column_number):
        # Convert an Excel column number into the corresponding column name.
        if column_number <= 0:
            raise ValueError("Column numbers must be 1 or greater")
        letters = []
        while column_number:
            column_number, remainder = divmod(column_number - 1, 26)
            letters.append(chr(65 + remainder))
        return ''.join(reversed(letters))


