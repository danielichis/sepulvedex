 
#-Includes--------------------------------------------------------------
import pandas as pd
import shutil
from datetime import *
from openpyxl import load_workbook
from copy import copy
import os

ruta_plantilla='C:\\Users\\Administrador.WIN-C8USBNGG6F4\\OneDrive - industrias venado\\Documentos\\conciliacion\\'

def read_excel():
    name_newfile="CUADRO DIGITACION_CONFRONTACION 2022 NRS.xlsx"
    dest_wb = load_workbook(name_newfile)
    sheet1 = dest_wb['Datos']
    #dest_sheet.cell(row=4, column=3).value = "Saldo al " + cfdate #31/12/2023
    
    
    rows = sheet1.rows
    
    nrow=0
    dict_mayor ={}
    # Extrae Registro de Cuentas Cobradoras
    for row in rows:
        nrow+=1
        col_a="" #sheet1.cell(row=nrow, column=0).value
        col_b=sheet1.cell(row=nrow, column=1).value
        col_c=sheet1.cell(row=nrow, column=2).value
        col_d=sheet1.cell(row=nrow, column=3).value
        col_e=sheet1.cell(row=nrow, column=4).value
        col_f=sheet1.cell(row=nrow, column=5).value
        col_g=sheet1.cell(row=nrow, column=6).value    
    
        dict_mayor[nrow] = [col_a,col_b,col_c,col_d,col_e,col_f,col_g]
    print(dict_mayor)

read_excel()